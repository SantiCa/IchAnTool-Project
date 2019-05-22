

def OpenImg(root):
    import config
    from PIL import Image
    import tkinter as tk
    from tkinter import filedialog
    import numpy as np  #modulo numerico
    import matplotlib.pyplot as plt #modulo de graficos
    global State
    import os
    config.rutaenlace=filedialog.askopenfilename(initialdir = config.PathtoWfolder,title = "Select file (jpg)")
    print("path to the open image"+config.rutaenlace)
    if config.rutaenlace !=  "" and config.rutaenlace !=  "InitialBlankRute":
        Path=tk.Label(root, bg='lemon chiffon', text= str(config.count)+". The path of the open image is" + str(config.rutaenlace))   #Escribe la ruta a la imagen abierta, con numero
        Path.grid(row=4+config.count,column=0,sticky="nw")  #aparece debajo de la imagen abierta anterior
        config.img=Image.open(config.rutaenlace)


    if config.count>1 and config.rutaenlace !=  "":
        State.destroy()
    Size= tk.Label(root,bg='lemon chiffon', text="y " + str(config.img.size[1]) + ", x " + str(config.img.size[0]) + " pixeles") #indica el tamaño de la imagen abierta
    Size.grid(row=4+config.count,column=1,sticky="nw")
    State=tk.Label(root,bg='lemon chiffon', text="CURRENT",font=("Arial",11,"bold"),fg="red")
    State.grid(row=4+config.count,column=2,sticky="nw")
    plt.imshow(np.asarray(config.img))  #utiliza matplotlib para mostrar la imagen con ejes
    plt.title("imagen abierta en pseudocolor")
    plt.show(block=False)

    if config.count==1 and config.rutaenlace !=  "":
        print ("first time executed")
        tk.Label(root, bg='lemon chiffon',text="File Path",font=("Arial",10,"bold")).grid(row=4,column=0)
        tk.Label(root, bg='lemon chiffon',text="Size",font=("Arial",10,"bold")).grid(row=4,column=1)
        tk.Label(root, bg='lemon chiffon',text="State",font=("Arial",10,"bold")).grid(row=4,column=2)
    Namefill=(os.path.split(config.rutaenlace)[1])
    if Namefill.endswith('.jpg'):
        Namefill = Namefill[:-4]
    config.NameImg=Namefill


####################################################################################################################################################################################################################################################################

def CoreAnalysis(frame):
    import config
    import numpy as np
    import tkinter as tk
    from tkinter import messagebox
    if config.img == None:
        messagebox.showerror("Image not open!","open an image first")
        return
    else:
        imgarray=np.array(config.img)
        suma=porcentage(imgarray) #asigna a una variable lista suma() los porcentages de pixeles "negros" a lo largo del eje y de la imagen
        excelcreate(suma,frame)
############################################################################################################################################################################################################################################
def porcentage(imgarray):
    b=[]    #loop que define una lista con la suma de los valore smayores de 100 por filas y despues devuelve el porcentage
             #pixeles que dan este valor
    n=[]
    for i in imgarray:
        b.append(n)
        n=[]
        for j in i:
            if j < 5:
                n.append(1)
    suma=[]
    for i in b:
        suma.append(sum(i)*100/imgarray.shape[1])
    print("Sum of  pixels done")
    return suma
##############################################################################################################################################################################################################################################
#0
def excelcreate(A,frame):
####1
    import config
    import datetime
    import tkinter as tk
    from tkinter import filedialog
    from openpyxl import Workbook
    import os


######2 (ecelcreate)
    def sheeteer(A,largoimg,Inicio): #array con el analysis realizado
      import datetime
      from tkinter import filedialog
      from openpyxl import Workbook
########3 (sheetcreator)
      def sheetcreator(A,PixDato,largoimg,Inicio,d): #funcion que crea las columnas que necesitamos en el excel #PixDato= indica el numero de pixeles (de filas qe queremos considerar para cada punto creado por el programa
        from openpyxl import Workbook
        ws = wb.create_sheet(str(PixDato) + " pixels mean") # insert at the end (default)
        a=0
        ws.cell(row=1, column=1).value="mean bioturbation x " + str(PixDato) +" pixeles"
        ws.cell(row=1, column=2).value="pixel"
        ws.cell(row=1, column=3).value="cm"
        ws.cell(row=1, column=4).value="BI (Bioturbation index), after after Reineck 1963, and Taylor and Goldring 1993"
        a=1
        i=0
        r=1
        c=0
        for x in A: #crea un loop for para la asignacion de valores necesarios
            c=c+x
            a=a+1
            i=i+1
            if i==PixDato:
                p=c/PixDato
                r=r+1
                u=((a-(float(PixDato/2))) * float(largoimg) / d)
                ws.cell(row=r, column=1).value=float(p)
                ws.cell(row=r, column=2).value=float(a-(PixDato/2))
                ws.cell(row=r, column=3).value=float(u+float(Inicio))
                if p<1:
                    aux=0
                elif p>=1 and p<5:
                    aux=1
                elif p>=5 and p<31:
                    aux=2
                elif p>=31 and p<61:
                    aux=3
                elif p>=61 and p<91:
                    aux=4
                elif p>=91 and p<100:
                    aux=5
                elif p>=100:
                    aux=6
                ws.cell(row=r, column=4).value=float(aux)
                c=0
                i=0
########3 (sheetcreator)
        print("sheet for "+str(PixDato)+" pixel interval created")
######2 (sheeter)
      wb =Workbook()
      WS0=wb.active
      WS0.title="Basic information"
      WS0.cell(row=1, column=1).value="obtained data"
      WS0.cell(row=1, column=2).value= "Date"
      WS0.cell(row=1, column=3).value=str(datetime.date.today())
      WS0.cell(row=2, column=1).value=" Size "
      WS0.cell(row=2, column=2).value=str(config.img.size[1])
      WS0.cell(row=2, column=3).value="Hight (pixels)"
      WS0.cell(row=2, column=4).value=str(config.img.size[0])
      WS0.cell(row=2, column=5).value="Width (pixels)"
      d=len(A)
      sheetcreator(A,1,largoimg,Inicio,d)
      sheetcreator(A,10,largoimg,Inicio,d)
      sheetcreator(A,50,largoimg,Inicio,d)
      sheetcreator(A,100,largoimg,Inicio,d)
      sheetcreator(A,200,largoimg,Inicio,d)
      print("Excel workbook generated")
      return wb

####1 (Excelcreator)
    L4=tk.Label(frame,bg="gold",text="Introduzca sus datos")
    L4.pack(fill=tk.X)
    Inicio=tk.IntVar()
    Fin=tk.IntVar()
    NameArchive=tk.StringVar()
    L1=tk.Label(frame, text='Depth of the Top of the core image')
    L1.pack(fill=tk.X)
    Entrybox1=tk. Entry(frame,textvariable=Inicio)
    Entrybox1.pack(fill=tk.X)
    L2=tk.Label(frame,text='Depth of the Bottom of the core image')
    L2.pack(fill=tk.X)
    Entrybox2=tk.Entry(frame, textvariable=Fin)
    Entrybox2.pack(fill=tk.X)
    L3=tk.Label(frame,text='Name of the file which will be generated (it is an excel file)')
    L3.pack(fill=tk.X)
    Entrybox3=tk.Entry(frame, textvariable=NameArchive)
    Entrybox3.pack(fill=tk.X)
    #### creacion de el frame dinamico de opciones de nombre

    OptionsFrame=tk.Frame(frame,bg='mint cream' )
    OptionsFrame.pack()
    IsNameImg=False

    OptionsFrame.grid_columnconfigure(0, weight=1)
    OptionsFrame.grid_columnconfigure(4, weight=1)
    #Obtencion nombre imagen abierta


########2 (autoassign)

    def autoasign():
        global B5
        global IsNameImg
        global L
        try:
            L.destroy()
        except (NameError, AttributeError):
            pass
        Entrybox3.config(state=tk.DISABLED)
        B5=tk.Button(OptionsFrame,text="Undo image name asign",command=undo)
        B5.grid(row= 0,column=2,sticky="ew" )
        L=tk.Label(OptionsFrame, text=config.NameImg)
        L.grid(row= 1,column=0, columnspan= 4,sticky="we")
        IsNameImg=True
########2 (undo)
    def undo():
        global IsNameImg
        global L
        Entrybox3.config(state=tk.NORMAL)
        B5.destroy()
        L.destroy()
        del L
        IsNameImg=False

    B4=tk.Button(OptionsFrame,text="Asing the same as image",command=lambda: autoasign())
    B4.grid(row= 0,column=1,sticky="ew")
########2 (getdata)
    def getdata(Inicio,Fin,NameArchive,frame):
        from tkinter import messagebox
        global IsNameImg
        Entrybox1.config(state=tk.DISABLED)
        Entrybox2.config(state=tk.DISABLED)
        Entrybox3.config(state=tk.DISABLED)
        B2.config(state=tk.DISABLED)
        B4.config(state=tk.DISABLED)
        B5.config(state=tk.DISABLED)
        Inicio= float(Inicio.get())
        Fin= float(Fin.get())
        NameArchive= str(NameArchive.get())
        largoimg=Fin-Inicio
        Wb=sheeteer(A,largoimg,Inicio) #paso importante, aqui se crea el wornook
        carpeta=filedialog.askdirectory(initialdir = config.PathtoWfolder,title ="select where to save the CSV archive")
        if carpeta=="":
            messagebox.showerror("No folder selected","A folder must be selected")
            Entrybox1.config(state=tk.NORMAL)
            Entrybox2.config(state=tk.NORMAL)
            Entrybox3.config(state=tk.NORMAL)
            B2.config(state=tk.NORMAL)
            B4.config(state=tk.NORMAL)
            B5.config(state=tk.NORMAL)
        SaveFrame=tk.Frame(frame,bg='lavender')
        SaveFrame.pack()
        tk.Label(SaveFrame,text="Save the Excel",bg="gold").pack(fill=tk.X)
        tk.Label(SaveFrame,text="Analisis parameters: Top= "+str(Inicio)+", Bottom= "+str(Fin)+", Core image vertical lengh= "+ str(largoimg),bg='lavender').pack()
        tk.Label(SaveFrame,text="Press the button to save in: "+ carpeta +" or 'back' to change the parameters or the directory",bg='lavender').pack()
        Name=""
        if IsNameImg==True:
            Name=config.NameImg
        else:
            Name=NameArchive
        tk.Label(SaveFrame,text= "The name of the archive will be: "+ Name +".xlsx",bg='lavender').pack()
############3 (save)
        def Save(Name,carpeta,Wb,SaveFrame):
            import os
            Wb.save(carpeta+"/"+ Name +".xlsx")
            tk.Label(SaveFrame,text="Saved succesfully in: "+ carpeta+"/"+ Name +".xlsx", font=("arial",11,"bold"),fg="red",bg='lavender').pack()
            B1.config(state=tk.DISABLED)
            os.startfile(carpeta +"/"+ Name + ".xlsx")
            Entrybox1.destroy()
            Entrybox2.destroy()
            Entrybox3.destroy()
            B2.destroy()
            L1.destroy()
            L2.destroy()
            L3.destroy()
            L4.destroy()
            OptionsFrame.destroy()
            print("Saved")
############3 (back)
        def back(frame):
            Entrybox1.destroy()
            Entrybox2.destroy()
            Entrybox3.destroy()
            B2.destroy()
            L1.destroy()
            L2.destroy()
            L3.destroy()
            L4.destroy()
            OptionsFrame.destroy()
            SaveFrame.destroy()
            CoreAnalysis(Rightframe)
########2 (getdata)
        B1=tk.Button(SaveFrame,text="Save",command = lambda: Save(Name,carpeta,Wb,SaveFrame))
        B1.pack()
        B3=tk.Button(SaveFrame,text="back",command = lambda: back(frame))
        B3.pack()

####1 (Excelcreator)
    B2=tk.Button(frame, text="Analyze and generate Excel File",command = lambda: getdata(Inicio,Fin,NameArchive,frame))
    B2.pack()
################################################################################
def Fullcreator(dirlist,window,ruta):
######2 (sheeter)
    import numpy as np
    from PIL import Image
    import os
    def sheeteer2(A,largoimg,Inicio,imagen): #array con el analysis realizado
      import datetime
      from tkinter import filedialog
      from openpyxl import Workbook

##########3 (sheetcreator2)
      def sheetcreator2(A,PixDato,largoimg,Inicio,d,imagen): #funcion que crea las columnas que necesitamos en el excel #PixDato= indica el numero de pixeles (de filas qe queremos considerar para cada punto creado por el programa
          from openpyxl import Workbook
          ws = wb.create_sheet(str(PixDato) + " pixels mean") # insert at the end (default)
          a=0


          ws.cell(row=1, column=1).value="pixel"
          ws.cell(row=1, column=2).value="cm"
          ws.cell(row=1, column=3).value="mean bioturbation x " + str(PixDato) +" pixeles"
          ws.cell(row=1, column=4).value="BI (Bioturbation index), after after Reineck 1963, and Taylor and Goldring 1993"

          a=1
          i=0
          r=1
          c=0
          for x in A: #crea un loop for para la asignacion de valores necesarios
              c=c+x
              a=a+1
              i=i+1
              if i==PixDato:
                p=c/PixDato
                r=r+1
                u=((a-(float(PixDato/2))) * float(largoimg) / d)

                ws.cell(row=r, column=1).value=float(a-(PixDato/2))
                ws.cell(row=r, column=2).value=float(u+float(Inicio))
                ws.cell(row=r, column=3).value=float(p)
                if p<1:
                    aux=0
                elif p>=1 and p<5:
                    aux=1
                elif p>=5 and p<31:
                    aux=2
                elif p>=31 and p<61:
                    aux=3
                elif p>=61 and p<91:
                    aux=4
                elif p>=91 and p<100:
                    aux=5
                elif p>=100:
                    aux=6
                ws.cell(row=r, column=4).value=float(aux)
                c=0
                i=0

######2 (sheeter)
      wb =Workbook()
      WS0=wb.active
      WS0.title="Basic information"
      WS0.cell(row=1, column=1).value="obtained data"
      WS0.cell(row=1, column=2).value= "Date"
      WS0.cell(row=1, column=3).value=str(datetime.date.today())
      WS0.cell(row=2, column=1).value=" Size "
      WS0.cell(row=2, column=2).value=str(imagen.size[1])
      WS0.cell(row=2, column=3).value="Hight (pixels)"
      WS0.cell(row=2, column=4).value=str(imagen.size[0])
      WS0.cell(row=2, column=5).value="Width (pixels)"
      d=len(A)
      sheetcreator2(A,1,largoimg2,Inicio2,d,imagen)
      sheetcreator2(A,10,largoimg2,Inicio2,d,imagen)
      sheetcreator2(A,50,largoimg2,Inicio2,d,imagen)
      sheetcreator2(A,100,largoimg2,Inicio2,d,imagen)
      sheetcreator2(A,200,largoimg2,Inicio2,d,imagen)
      print("Excel workbook generated")
      return wb

#############################################


    cont=0
    for i in dirlist:   ##bucle principal
        cont=cont+1
        imagen=Image.open(ruta+"/"+i)
        imagenArray=np.array(imagen)
        suma1=porcentage(imagenArray)
        Name=(os.path.split(i)[1])
        largoimg2 = float(Name[-11:-4])-float(Name[:7])
        Inicio2=float(Name[:7])
        wb=sheeteer2(suma1,largoimg2,Inicio2,imagen)
        carpeta2=ruta
        wb.save(carpeta2+"/tramo_"+str(cont)+ str(i[:-4])+".xlsx")
        carpeta2=dirlist
def CreateFullFolder(root):
    import config
    import datetime
    import tkinter as tk
    from tkinter import filedialog
    from openpyxl import Workbook
    import os


    window=tk.Toplevel()
    window.title('Full Folder Analysis alfa 1.0')
    window.wm_iconbitmap("IconWindow.ico")
    window.wm_title("IC")
    window.attributes("-topmost", True)
    L1=tk.Label(window,text="Select the folder where the images are located",bg="gold").grid(row=0 ,column=1,columnspan=3,sticky="ew")
    ruta=filedialog.askdirectory(title='Please select a directory where images wich will be merged are located ')
    dirlist=os.listdir(ruta)
    L2=tk.Label(window,text="the selected carpet is: ").grid(row=2)
    L3=tk.Label(window,text=ruta).grid(row=3)

    B1=tk.Button(window, text="GO!", command=lambda:Fullcreator(dirlist,window,ruta)).grid(row=4)
